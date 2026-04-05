import tempfile, os, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def _thin():
    s = Side(style="thin", color="8EAACC")
    return Border(top=s, bottom=s, left=s, right=s)

def _header_fill():
    return PatternFill("solid", start_color="DAEEF3")

def build_payslip(data: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "給料明細"

    for c, w in [("A",12),("B",14),("C",14),("D",14),("E",14),("F",14),("G",14),("H",6)]:
        ws.column_dimensions[c].width = w
    for r in range(1, 22):
        ws.row_dimensions[r].height = 20

    # ヘッダー
    ws["A1"] = "株式会社　坂野建築"
    ws["A1"].font = Font(name="メイリオ", bold=True)
    ws["A2"] = f"氏名　{data.get('name','')}"
    ws["A3"] = f"社員番号　{data.get('emp_no','')}"

    period = data.get("period","")
    m = re.match(r"(\d{4})年(\d{1,2})月", period)
    if m:
        rw = int(m.group(1)) - 2018
        mo = int(m.group(2))
        period_str = f"令和{rw}年{mo}月"
    else:
        period_str = period

    ws.merge_cells("D1:H1")
    ws["D1"] = f"{period_str}　給料支給明細書"
    ws["D1"].font = Font(name="メイリオ", bold=True, size=14)
    ws["D1"].alignment = Alignment(horizontal="center")

    def section(start_row, name, label_row, data_rows):
        ws.merge_cells(f"A{start_row}:H{start_row}")
        c = ws[f"A{start_row}"]
        c.value = name
        c.font = Font(bold=True, name="メイリオ")
        c.alignment = Alignment(horizontal="center")
        c.fill = PatternFill("solid", start_color="B8CCE4")
        c.border = _thin()

    # 勤怠
    section(5, "勤　怠", 6, [7])
    for col, label in enumerate(["就業日数","出勤日数","労働時間","休日出勤日数","有給消化日数"], 1):
        ws.cell(6, col, label).fill = _header_fill()
        ws.cell(6, col).font = Font(bold=True, size=9)
        ws.cell(6, col).border = _thin()
        ws.cell(6, col).alignment = Alignment(horizontal="center")

    keys = ["work_days","attend_days","work_hours","holiday_work","paid_leave"]
    for col, key in enumerate(keys, 1):
        ws.cell(7, col, data.get(key, "")).border = _thin()
        ws.cell(7, col).alignment = Alignment(horizontal="center")

    for col, label in enumerate(["残業","休日出勤"], 1):
        ws.cell(8, col, label).fill = _header_fill()
        ws.cell(8, col).font = Font(bold=True, size=9)
        ws.cell(8, col).border = _thin()
        ws.cell(8, col).alignment = Alignment(horizontal="center")
    ws.cell(9, 1, data.get("overtime","")).border = _thin()
    ws.cell(9, 2, data.get("holiday_work","")).border = _thin()

    # 支給
    section(11, "支　給", 12, [13])
    for col, label in enumerate(["基本給","残業手当","深夜勤務手当","役員報酬"], 1):
        ws.cell(12, col, label).fill = _header_fill()
        ws.cell(12, col).font = Font(bold=True, size=9)
        ws.cell(12, col).border = _thin()
        ws.cell(12, col).alignment = Alignment(horizontal="center")
    for col, key in enumerate(["base_salary","overtime_pay","night_pay","director_pay"], 1):
        c = ws.cell(13, col, data.get(key, 0))
        c.number_format = "#,##0"
        c.border = _thin()
        c.alignment = Alignment(horizontal="right")

    # 控除
    section(15, "控　除", 16, [17])
    for col, label in enumerate(["厚生年金保険","雇用保険","健康保険","所得税","家賃"], 1):
        ws.cell(16, col, label).fill = _header_fill()
        ws.cell(16, col).font = Font(bold=True, size=9)
        ws.cell(16, col).border = _thin()
        ws.cell(16, col).alignment = Alignment(horizontal="center")
    for col, key in enumerate(["pension","employment_ins","health_ins","income_tax","rent"], 1):
        c = ws.cell(17, col, data.get(key, 0))
        c.number_format = "#,##0"
        c.border = _thin()
        c.alignment = Alignment(horizontal="right")

    # 合計
    section(19, "合　計", 20, [21])
    base = data.get("base_salary",0)+data.get("overtime_pay",0)+data.get("night_pay",0)+data.get("director_pay",0)
    ded  = data.get("pension",0)+data.get("employment_ins",0)+data.get("health_ins",0)+data.get("income_tax",0)+data.get("rent",0)
    net  = base - ded
    for col, label, val in [(1,"総支給額",base),(3,"総控除額",ded),(5,"差引支給額",net)]:
        ws.cell(20, col, label).fill = _header_fill()
        ws.cell(20, col).font = Font(bold=True, size=9)
        ws.cell(20, col).border = _thin()
        ws.merge_cells(start_row=20, start_column=col, end_row=20, end_column=col+1)
        c = ws.cell(21, col, val)
        c.number_format = "#,##0"
        c.font = Font(bold=True)
        c.border = _thin()
        ws.merge_cells(start_row=21, start_column=col, end_row=21, end_column=col+1)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r'[\\/:*?"<>|]', '', data.get("name","社員"))
    path = os.path.join(tempfile.gettempdir(), f"給料明細_{safe}_{timestamp}.xlsx")
    wb.save(path)
    return path
