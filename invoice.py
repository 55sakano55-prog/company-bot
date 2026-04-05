import tempfile, os, re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

def _thin():
    s = Side(style="thin", color="000000")
    return Border(top=s, bottom=s, left=s, right=s)

def build_invoice(data: dict) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "請求書"

    # 列幅
    for c, w in [("A",4),("B",4),("C",8),("D",8),("E",8),("F",8),
                 ("G",8),("H",8),("I",8),("J",8),("K",8),("L",8)]:
        ws.column_dimensions[c].width = w

    # タイトル
    ws.merge_cells("A1:L1")
    ws["A1"] = "請　求　書"
    ws["A1"].font = Font(name="メイリオ", bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    # 発行日
    date_str = data.get("date", "")
    m = re.match(r"(\d{4})年(\d{1,2})月(\d{1,2})日", date_str)
    if m:
        y,mo,d = int(m.group(1))-2018, int(m.group(2)), int(m.group(3))
        ws.merge_cells("H2:L2")
        ws["H2"] = f"令和{y}年{mo}月{d}日"
        ws["H2"].alignment = Alignment(horizontal="right")

    # 宛先
    ws.merge_cells("A3:F3")
    ws["A3"] = data.get("client","") + "　御中"
    ws["A3"].font = Font(name="メイリオ", bold=True, size=12)

    # 発行者情報
    for r, txt in [(4,"〒361-0067"),(5,"埼玉県行田市下池守207-14"),
                   (6,"株式会社坂野建築"),(7,"TEL: 080-8840-6906"),(8,"担当: 坂野 陽祐")]:
        ws.merge_cells(f"H{r}:L{r}")
        ws[f"H{r}"] = txt
        ws[f"H{r}"].alignment = Alignment(horizontal="right")

    ws["A5"] = "下記の通り御請求申し上げます。"
    ws.merge_cells("A6:F6")
    ws["A6"] = f"現場名：{data.get('site','')}"
    ws.merge_cells("A7:F7")
    ws["A7"] = f"支払期限：{data.get('due_date','')}"
    ws.merge_cells("A8:F8")
    ws["A8"] = "支払方法：銀行振込"

    # 明細ヘッダー
    ws.merge_cells("A10:L10")
    ws["A10"] = "御請求金額（税込）"
    ws["A10"].font = Font(bold=True, size=12)

    row = 12
    for col, label, width in [("A","項番",2),("B","工事名",6),("I","数量",1),("J","単価",1),("K","金額",2)]:
        ws.merge_cells(f"{col}{row}:{get_column_letter(ws[col+str(row)].column+width-1)}{row}")
        ws[f"{col}{row}"] = label
        ws[f"{col}{row}"].font = Font(bold=True)
        ws[f"{col}{row}"].fill = PatternFill("solid", start_color="DEEAF1")
        ws[f"{col}{row}"].alignment = Alignment(horizontal="center")
        ws[f"{col}{row}"].border = _thin()

    # 明細行
    subtotal = 0
    items = data.get("items", [])
    for i, item in enumerate(items[:10]):
        r = 13 + i
        ws[f"A{r}"] = i + 1
        ws.merge_cells(f"B{r}:H{r}")
        ws[f"B{r}"] = item["name"]
        ws[f"I{r}"] = item["qty"]
        ws[f"J{r}"] = item["price"]
        amt = item["qty"] * item["price"]
        ws.merge_cells(f"K{r}:L{r}")
        ws[f"K{r}"] = amt
        ws[f"K{r}"].number_format = "#,##0"
        subtotal += amt
        for cell in [ws[f"A{r}"], ws[f"B{r}"], ws[f"I{r}"], ws[f"J{r}"], ws[f"K{r}"]]:
            cell.border = _thin()

    # 集計
    tax = int(subtotal * 0.1)
    total = subtotal + tax
    for r, label, val in [(24,"小　計",subtotal),(25,"消費税",tax),(26,"合　計",total)]:
        ws.merge_cells(f"A{r}:J{r}")
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True)
        ws[f"A{r}"].border = _thin()
        ws.merge_cells(f"K{r}:L{r}")
        ws[f"K{r}"] = val
        ws[f"K{r}"].number_format = "#,##0"
        ws[f"K{r}"].border = _thin()

    # 御請求金額
    ws["E10"] = total
    ws["E10"].number_format = "#,##0"
    ws["E10"].font = Font(bold=True, size=14)

    # 備考
    ws.merge_cells("A28:L29")
    ws["A28"] = f"備考：{data.get('note','')}"

    # 振込先
    ws.merge_cells("A31:L31")
    ws["A31"] = "振込先：埼玉りそな銀行 行田支店　(普通) No.4669802　カ）サカノケンチク"

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r'[\\/:*?"<>|]', '', data.get("client","請求先"))
    path = os.path.join(tempfile.gettempdir(), f"請求書_{safe}_{timestamp}.xlsx")
    wb.save(path)
    return path
